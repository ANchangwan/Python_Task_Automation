from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("yourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"]

print(wb.sheetnames)

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"



wb.save("practice.xlsx")