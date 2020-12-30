from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 90:
        print(row[0].value,"번 학생의 영어 성적", end = " ")
        print(row[1].value)

for row in ws.iter_rows(max_row = 1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_hi.xlsx")